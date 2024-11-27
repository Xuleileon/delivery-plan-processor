import logging
from pathlib import Path
from main import process_delivery_plan

def test_local_excel():
    """测试处理本地Excel文件"""
    # 设置日志级别
    logging.basicConfig(level=logging.INFO)
    
    # 测试文件路径
    test_file = Path(__file__).parent / 'test_data' / 'test.xlsx'
    
    # 确保测试目录存在
    test_file.parent.mkdir(parents=True, exist_ok=True)
    
    # 如果测试文件不存在，创建一个简单的Excel文件
    if not test_file.exists():
        import openpyxl
        wb = openpyxl.Workbook()
        
        # 创建常规产品工作表
        ws1 = wb.active
        ws1.title = "常规产品"
        # 设置列名
        columns = ["sku编码", "商品名称", "规格", "到货批次-1", "到货数量-1", "到货批次-2", "到货数量-2", "到货批次-3", "到货数量-3", "到货批次-4", "到货数量-4", "到货批次-5", "到货数量-5"]
        for col, name in enumerate(columns, 1):
            ws1.cell(row=1, column=col, value=name)
        # 添加测试数据
        import datetime
        today = datetime.datetime.now()
        ws1.append([
            "001", "测试产品1", "颜色:红色,尺码:L", 
            today + datetime.timedelta(days=1), 100,
            today + datetime.timedelta(days=2), 50,
            today + datetime.timedelta(days=3), 30,
            today + datetime.timedelta(days=4), 20,
            today + datetime.timedelta(days=5), 10
        ])
        ws1.append([
            "002", "测试产品2", "颜色:蓝色,尺码:XL",
            today + datetime.timedelta(days=1), 200,
            today + datetime.timedelta(days=2), 100,
            today + datetime.timedelta(days=3), 50,
            today + datetime.timedelta(days=4), 30,
            today + datetime.timedelta(days=5), 20
        ])
        
        # 创建S级产品工作表
        ws2 = wb.create_sheet("S级产品")
        # 设置列名
        columns = ["sku编码", "商品名称"] + [f"2024/11/{i}" for i in range(14, 32)] + [f"2024/12/{i}" for i in range(1, 32)] + [f"2025/1/{i}" for i in range(1, 32)]
        for col, name in enumerate(columns, 1):
            ws2.cell(row=1, column=col, value=name)
        # 添加测试数据
        row_data = ["003", "测试产品3"] + [100] * (len(columns) - 2)
        ws2.append(row_data)
        row_data = ["004", "测试产品4"] + [200] * (len(columns) - 2)
        ws2.append(row_data)
        
        # 创建汇总工作表
        ws3 = wb.create_sheet("汇总")
        # 设置列名
        columns = ["sku编码", "sku名称", "总数量"]
        for col, name in enumerate(columns, 1):
            ws3.cell(row=1, column=col, value=name)
        
        # 保存文件
        wb.save(test_file)
    
    try:
        # 处理文件
        result = process_delivery_plan(str(test_file))
        
        # 检查结果
        assert result['success'], f"处理失败: {result['message']}"
        print("本地Excel测试成功！")
        print(f"预处理文件: {result['data']['preprocessed_file']}")
        print(f"转换后文件: {result['data']['transformed_file']}")
        print(f"最终文件: {result['data']['final_file']}")
        
    except Exception as e:
        print(f"本地Excel测试失败: {str(e)}")
        raise

def test_feishu():
    """测试从飞书获取数据"""
    # 设置日志级别
    logging.basicConfig(level=logging.INFO)
    
    try:
        # 处理飞书数据
        result = process_delivery_plan(None)  # 传入None表示使用飞书数据
        
        # 检查结果
        assert result['success'], f"处理失败: {result['message']}"
        print("飞书数据测试成功！")
        print(f"预处理文件: {result['data']['preprocessed_file']}")
        print(f"转换后文件: {result['data']['transformed_file']}")
        print(f"最终文件: {result['data']['final_file']}")
        
    except Exception as e:
        print(f"飞书数据测试失败: {str(e)}")
        raise

if __name__ == "__main__":
    print("=== 测试飞书数据处理 ===")
    test_feishu()
    print("\n=== 测试本地Excel处理 ===")
    test_local_excel()
