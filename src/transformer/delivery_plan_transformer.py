import logging
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from src.core.config import config
from src.core.exceptions import FileOperationError, DataTransformError
from src.utils.excel_utils import generate_output_path
from datetime import datetime, timedelta, time
import re

class DeliveryPlanTransformer:
    """到货计划转换器"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        
    def transform(self, input_file: Path) -> Path:
        """
        转换到货计划
        
        Args:
            input_file: 输入文件路径
            
        Returns:
            转换后的文件路径
            
        Raises:
            FileOperationError: 文件操作失败时抛出
            DataTransformError: 数据转换失败时抛出
        """
        try:
            # 检查文件是否存在
            if not input_file.exists():
                raise FileOperationError(f"输入文件不存在: {input_file}")
                
            # 读取数据
            sheets = self._read_sheets(input_file)
            
            # 转换格式
            transformed_df = self._transform_format(sheets)
            
            # 保存结果
            output_path = generate_output_path(
                input_file,
                Path(config.output_config['directory']),
                config.output_config['suffixes']['transform']
            )
            self._save_result(transformed_df, output_path)
            
            return output_path
            
        except Exception as e:
            self.logger.error(f"转换失败: {str(e)}", exc_info=True)
            if isinstance(e, (FileOperationError, DataTransformError)):
                raise
            raise DataTransformError(f"转换失败: {str(e)}")
            
    def _read_sheets(self, input_file: Path) -> dict:
        """读取所有工作表"""
        try:
            # 读取常规产品和S级产品工作表
            sheets_data = {}
            sheets_data['regular'] = pd.read_excel(input_file, sheet_name='常规产品')
            sheets_data['s_level'] = pd.read_excel(input_file, sheet_name='S级产品')
            return sheets_data
            
        except Exception as e:
            raise DataTransformError(f"读取工作表失败: {str(e)}")
        
    def _transform_format(self, sheets):
        """转换数据格式"""
        try:
            # 合并SKU
            regular_df = sheets['regular']
            s_level_df = sheets['s_level']
            
            # 打印原始数据信息
            regular_total = 0
            self.logger.info("常规产品数据:")
            for i in range(1, 6):
                date_col = f'到货批次-{i}'
                qty_col = f'到货数量-{i}'
                if date_col in regular_df.columns and qty_col in regular_df.columns:
                    total_qty = regular_df[qty_col].sum()
                    regular_total += total_qty
                    self.logger.info(f"{qty_col}总和: {total_qty}")
            self.logger.info(f"常规产品总和: {regular_total}")
                    
            s_level_total = 0
            self.logger.info("S级产品数据:")
            for col in s_level_df.columns:
                try:
                    if isinstance(col, datetime) or pd.to_datetime(col, errors='coerce') is not pd.NaT:
                        total_qty = s_level_df[col].sum()
                        s_level_total += total_qty
                        self.logger.info(f"{col}总和: {total_qty}")
                except (ValueError, TypeError):
                    continue
            self.logger.info(f"S级产品总和: {s_level_total}")
            self.logger.info(f"理论总和: {regular_total + s_level_total}")
            
            # 处理常规产品数据
            regular_data = {}
            for _, row in regular_df.iterrows():
                # 尝试不同的列名
                sku = None
                for col_name in ['sku编码', 'SKU编码', 'sku_no']:
                    if col_name in row.index:
                        sku = str(row[col_name]).strip()
                        break
                
                if not sku or pd.isna(sku) or sku == '0' or sku == '':
                    continue
                    
                if sku not in regular_data:
                    regular_data[sku] = {
                        'color': None,
                        'size': None,
                        'dates': {}
                    }
                    
                # 从规格提取颜色和尺码
                spec_col = None
                for col_name in ['规格', '商品规格']:
                    if col_name in row.index:
                        spec_col = col_name
                        break
                
                if spec_col and pd.notna(row[spec_col]):
                    color_match = re.search(r'颜色[:：\s]*([^,，\s]+)', str(row[spec_col]))
                    size_match = re.search(r'尺码[:：\s]*([^,，\s]+)', str(row[spec_col]))
                    regular_data[sku]['color'] = color_match.group(1) if color_match else ''
                    regular_data[sku]['size'] = size_match.group(1) if size_match else ''
                
                # 处理日期和数量
                for i in range(1, 6):
                    date_col = f'到货批次-{i}'
                    qty_col = f'到货数量-{i}'
                    if date_col in row.index and qty_col in row.index:
                        try:
                            date = row[date_col]
                            qty = row[qty_col]
                            if pd.notna(date) and pd.notna(qty):
                                # 尝试转换日期
                                try:
                                    if isinstance(date, datetime):
                                        date = date.date()
                                    elif isinstance(date, time):
                                        # 如果是time类型，跳过这个值
                                        continue
                                    elif isinstance(date, str):
                                        # 尝试多种日期格式
                                        for fmt in config.date_config['input_formats']:
                                            try:
                                                date = datetime.strptime(date, fmt).date()
                                                break
                                            except ValueError:
                                                continue
                                        else:
                                            raise ValueError(f"无效的日期格式: {date}")
                                    else:
                                        raise ValueError(f"无效的日期类型: {type(date)}")
                                except Exception as e:
                                    raise DataTransformError(f"日期转换失败: {str(e)}")
                                
                                # 如果日期已存在，累加数量
                                if date in regular_data[sku]['dates']:
                                    regular_data[sku]['dates'][date] += float(qty)
                                else:
                                    regular_data[sku]['dates'][date] = float(qty)
                        except (ValueError, TypeError) as e:
                            self.logger.warning(f"处理常规产品数据时出错: SKU={sku}, 日期={date_col}, 数量={qty_col}, 错误={str(e)}")
                            continue
            
            # 处理S级产品数据
            s_level_data = {}
            for _, row in s_level_df.iterrows():
                # 尝试不同的列名
                sku = None
                for col_name in ['sku编码', 'SKU编码', 'sku_no']:
                    if col_name in row.index:
                        sku = str(row[col_name]).strip()
                        break
                
                if not sku or pd.isna(sku) or sku == '0' or sku == '':
                    continue
                    
                if sku not in s_level_data:
                    s_level_data[sku] = {
                        'color': None,
                        'size': None,
                        'dates': {}
                    }
                    
                # 从规格提取颜色和尺码
                spec_col = None
                for col_name in ['规格', '商品规格']:
                    if col_name in row.index:
                        spec_col = col_name
                        break
                
                if spec_col and pd.notna(row[spec_col]):
                    color_match = re.search(r'颜色[:：\s]*([^,，\s]+)', str(row[spec_col]))
                    size_match = re.search(r'尺码[:：\s]*([^,，\s]+)', str(row[spec_col]))
                    s_level_data[sku]['color'] = color_match.group(1) if color_match else ''
                    s_level_data[sku]['size'] = size_match.group(1) if size_match else ''
                
                # 处理日期和数量
                for col in s_level_df.columns:
                    try:
                        # 尝试转换日期列
                        try:
                            if isinstance(col, datetime):
                                date = col.date()
                            elif isinstance(col, time):
                                # 如果是time类型，跳过这个列
                                continue
                            elif isinstance(col, str):
                                # 尝试多种日期格式
                                for fmt in config.date_config['input_formats']:
                                    try:
                                        date = datetime.strptime(col, fmt).date()
                                        break
                                    except ValueError:
                                        continue
                                else:
                                    continue  # 不是日期列，跳过
                            else:
                                continue  # 不是日期列，跳过
                        except Exception:
                            continue  # 日期转换失败，跳过
                            
                        qty = row[col]
                        if pd.notna(qty):
                            # 如果日期已存在，累加数量
                            if date in s_level_data[sku]['dates']:
                                s_level_data[sku]['dates'][date] += float(qty)
                            else:
                                s_level_data[sku]['dates'][date] = float(qty)
                    except (ValueError, TypeError) as e:
                        self.logger.warning(f"处理S级产品数据时出错: SKU={sku}, 列={col}, 错误={str(e)}")
                        continue
            
            # 合并所有数据
            all_data = {}
            for sku, data in regular_data.items():
                if sku not in all_data:
                    all_data[sku] = {
                        'color': data['color'],
                        'size': data['size'],
                        'dates': data['dates'].copy()
                    }
                else:
                    for date, qty in data['dates'].items():
                        if date in all_data[sku]['dates']:
                            all_data[sku]['dates'][date] += qty
                        else:
                            all_data[sku]['dates'][date] = qty
                    
            for sku, data in s_level_data.items():
                if sku not in all_data:
                    all_data[sku] = {
                        'color': data['color'],
                        'size': data['size'],
                        'dates': data['dates'].copy()
                    }
                else:
                    for date, qty in data['dates'].items():
                        if date in all_data[sku]['dates']:
                            all_data[sku]['dates'][date] += qty
                        else:
                            all_data[sku]['dates'][date] = qty
            
            # 创建结果DataFrame
            today = datetime.now().date()
            yesterday = today - timedelta(days=1)
            
            result_data = []
            total_qty = 0  # 用于跟踪总数量
            
            for sku, data in all_data.items():
                row = {
                    'sku_no': sku,
                    'color': data['color'] or '',
                    'size': data['size'] or ''
                }
                
                # 添加60天的数据
                sku_total = 0  # 用于跟踪每个SKU的总数量
                for i in range(1, 61):
                    target_date = today + timedelta(days=i-1)
                    qty = float(data['dates'].get(target_date, 0))
                    row[f'day{i}'] = qty
                    sku_total += qty
                
                # 在最后添加dt列
                row['dt'] = yesterday.strftime(config.date_config['output_format'])
                
                total_qty += sku_total  # 累加到总数量
                self.logger.debug(f"SKU {sku} 总数量: {sku_total}")
                result_data.append(row)
            
            # 如果没有数据，创建一个空的DataFrame
            if not result_data:
                columns = ['sku_no', 'color', 'size'] + [f'day{i}' for i in range(1, 61)] + ['dt']
                result_df = pd.DataFrame(columns=columns)
            else:
                result_df = pd.DataFrame(result_data)
            
            # 确保所有数值列为float类型
            date_cols = [f'day{i}' for i in range(1, 61)]
            for col in date_cols:
                if col not in result_df.columns:
                    result_df[col] = 0.0
                else:
                    result_df[col] = pd.to_numeric(result_df[col], errors='coerce').fillna(0.0)
            
            # 打印数据总和
            total_sum = result_df[date_cols].sum().sum()
            day1_sum = result_df['day1'].sum() if 'day1' in result_df.columns else 0
            self.logger.info(f"转换后数据总和: {total_sum}")
            self.logger.info(f"转换后day1总和: {day1_sum}")
            
            # 打印每个SKU的数据
            for _, row in result_df.iterrows():
                sku = row['sku_no']
                sku_sum = sum(row[col] for col in date_cols if col in row)
                self.logger.debug(f"SKU {sku} 转换后总和: {sku_sum}")
            
            return result_df
            
        except Exception as e:
            self.logger.error(f"转换格式失败: {str(e)}", exc_info=True)
            if isinstance(e, DataTransformError):
                raise
            raise DataTransformError(f"转换格式失败: {str(e)}")
    
    def _save_result(self, df: pd.DataFrame, output_path: Path) -> None:
        """保存结果"""
        try:
            # 保存Excel文件
            df.to_excel(output_path, sheet_name='汇总', index=False)
            
            # 设置单元格格式
            wb = openpyxl.load_workbook(output_path)
            ws = wb['汇总']
            
            # 设置表头样式
            header_fill = PatternFill(start_color='1F6B3B', end_color='1F6B3B', fill_type='solid')
            for cell in ws[1]:
                cell.fill = header_fill
            
            wb.save(output_path)
            
            self.logger.info(f"已保存转换结果: {output_path}")
            
        except Exception as e:
            raise DataTransformError(f"保存结果失败: {str(e)}")