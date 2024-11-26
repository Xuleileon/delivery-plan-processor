from datetime import datetime
import openpyxl
from copy import copy
import pandas as pd
import os

def process_excel_file(input_file_path):
    try:
        print(f"尝试打开文件: {input_file_path}")  # 添加调试信息
        # 生成输出文件路径
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = os.path.splitext(os.path.basename(input_file_path))[0]  # 获取不带扩展名的文件名
        output_file = f"{filename}_{timestamp}.xlsx"
        output_path = os.path.join(os.path.dirname(input_file_path), output_file)
        
        # 加载原始 Excel 文件以保留所有格式和颜色
        wb = openpyxl.load_workbook(input_file_path)
        sheet1_wb = wb['常规产品']
        sheet2_wb = wb['S级产品']
        sheet3_wb = wb['汇总']

        # 重新加载数据表至 DataFrame 中进行处理
        sheet1 = pd.read_excel(input_file_path, sheet_name='常规产品')
        sheet2 = pd.read_excel(input_file_path, sheet_name='S级产品')
        sheet3 = pd.read_excel(input_file_path, sheet_name='汇总')

        # 收集所有日期
        all_dates = set()
        
        # 从sheet1收集日期
        for idx in range(1, 4):
            date_col = f'到货批次-{idx}'
            if date_col in sheet1.columns:
                dates = pd.to_datetime(sheet1[date_col], errors='coerce')
                valid_dates = dates.dropna().dt.strftime('%Y-%m-%d').unique()
                all_dates.update(valid_dates)

        # 从sheet2收集日期
        sheet2_dates = [col for col in sheet2.columns[3:] if pd.notna(pd.to_datetime(col, errors='coerce'))]
        all_dates.update([pd.to_datetime(date).strftime('%Y-%m-%d') for date in sheet2_dates])

        # 排序日期
        sorted_dates = sorted(all_dates, key=lambda x: pd.to_datetime(x))

        # 创建新的DataFrame，包含所有需要的列
        sheet3_cleaned = pd.DataFrame(columns=['sku编码', 'spu编码', '商品名称', '规格', '第一批时间', '第一批数量', '总交货量'] + sorted_dates)
        
        # 合并SKU
        unique_skus_new = pd.concat([
            sheet1['sku编码'].dropna(), 
            sheet2['sku编码'].dropna()
        ]).drop_duplicates().reset_index(drop=True)
        
        sheet3_cleaned['sku编码'] = unique_skus_new

        # 映射商品名称和规格
        sku_data_combined = pd.concat([
            sheet1[['sku编码', '商品名称', '规格']].dropna(subset=['sku编码']), 
            sheet2[['sku编码', '商品名称', '规格']].dropna(subset=['sku编码'])
        ], ignore_index=True).drop_duplicates(subset=['sku编码'])
        
        # 确保所需列存在
        required_cols = ['商品名称', '规格'] 
        for col in required_cols:
            if col not in sku_data_combined.columns:
                sku_data_combined[col] = None

        sku_mapping = sku_data_combined.set_index('sku编码')
        sheet3_cleaned['商品名称'] = sheet3_cleaned['sku编码'].map(sku_mapping['商品名称'])
        sheet3_cleaned['规格'] = sheet3_cleaned['sku编码'].map(sku_mapping['规格'])

        # 处理总交货量
        def get_total_quantity(row):
            # 获取所有日期列
            date_columns = [col for col in row.index if pd.notna(pd.to_datetime(col, errors='coerce'))]
            # 将数据转换为数值类型并计算总和
            values = pd.to_numeric(row[date_columns], errors='coerce')
            total = values.sum()
            
            # 添加验证输出
            if pd.notna(total):
                print(f"SKU: {row['sku编码']}, 各日期数量: {values.to_dict()}, 总和: {total}")
            
            return total if pd.notna(total) else 0

        # 处理sheet1的日期和数量
        sheet1_data = {}
        for idx in range(1, 6):
            date_col = f'到货批次-{idx}'
            qty_col = f'到货数量-{idx}'
            if date_col in sheet1.columns and qty_col in sheet1.columns:
                dates = pd.to_datetime(sheet1[date_col], errors='coerce')
                quantities = sheet1[qty_col]
                for sku, date, qty in zip(sheet1['sku编码'], dates, quantities):
                    if pd.notna(date) and pd.notna(qty):
                        date_str = date.strftime('%Y-%m-%d')
                        if sku not in sheet1_data:
                            sheet1_data[sku] = {}
                        if date_str in sheet1_data[sku]:
                            sheet1_data[sku][date_str] += qty
                        else:
                            sheet1_data[sku][date_str] = qty

        # 处理sheet2的日期和数量
        sheet2_data = {}
        for _, row in sheet2.iterrows():
            sku = row['sku编码']
            sheet2_data[sku] = {}
            for date_col in sheet2_dates:
                qty = row[date_col]
                if pd.notna(qty) and qty > 0:
                    date_str = pd.to_datetime(date_col).strftime('%Y-%m-%d')
                    sheet2_data[sku][date_str] = qty

        # 填充数据
        for idx, row in sheet3_cleaned.iterrows():
            sku = row['sku编码']
            # 填充sheet1的数据
            if sku in sheet1_data:
                for date, qty in sheet1_data[sku].items():
                    if date in sheet3_cleaned.columns:
                        current_qty = sheet3_cleaned.at[idx, date]
                        # 如果当前位置已经有数据，需要累加而不是覆盖
                        if pd.notna(current_qty):
                            sheet3_cleaned.at[idx, date] = current_qty + qty
                        else:
                            sheet3_cleaned.at[idx, date] = qty
            
            # 填充sheet2的数据
            if sku in sheet2_data:
                for date, qty in sheet2_data[sku].items():
                    if date in sheet3_cleaned.columns:
                        current_qty = sheet3_cleaned.at[idx, date]
                        # 如果当前位置已经有数据，需要累加而不是覆盖
                        if pd.notna(current_qty):
                            sheet3_cleaned.at[idx, date] = current_qty + qty
                        else:
                            sheet3_cleaned.at[idx, date] = qty

        # 填充数据后重新计算总交货量
        for idx, row in sheet3_cleaned.iterrows():
            sku = row['sku编码']
            # 填充sheet1和sheet2的数据（保持原有代码不变）
            if sku in sheet1_data:
                for date, qty in sheet1_data[sku].items():
                    if date in sheet3_cleaned.columns:
                        sheet3_cleaned.at[idx, date] = qty
            
            if sku in sheet2_data:
                for date, qty in sheet2_data[sku].items():
                    if date in sheet3_cleaned.columns:
                        sheet3_cleaned.at[idx, date] = qty

        # 重新计算总交货量
        sheet3_cleaned['总交货量'] = sheet3_cleaned.apply(get_total_quantity, axis=1)

        # 处理第一批数据
        def get_first_batch(sku):
            dates_qty = {}
            if sku in sheet1_data:
                dates_qty.update(sheet1_data[sku])
            if sku in sheet2_data:
                dates_qty.update(sheet2_data[sku])
            
            if dates_qty:
                first_date = min(dates_qty.keys())
                return pd.to_datetime(first_date).strftime('%Y-%m-%d'), dates_qty[first_date]
            return None, None

        # 填充第一批数据
        for idx, row in sheet3_cleaned.iterrows():
            first_date, first_qty = get_first_batch(row['sku编码'])
            if first_date and first_qty:
                sheet3_cleaned.at[idx, '第一批时间'] = first_date
                sheet3_cleaned.at[idx, '第一批数量'] = first_qty

        # 确保所有数值列都是数值类型
        numeric_columns = ['总交货量', '第一批数量'] + sorted_dates
        for col in numeric_columns:
            if col in sheet3_cleaned.columns:
                sheet3_cleaned[col] = pd.to_numeric(sheet3_cleaned[col], errors='coerce')

        # 替换 NA 值为 None
        sheet3_cleaned = sheet3_cleaned.where(pd.notna(sheet3_cleaned), None)

        # 清除现有数据
        for row in sheet3_wb.iter_rows(min_row=1, max_row=sheet3_wb.max_row):
            for cell in row:
                cell.value = None

        # 保存列宽
        column_widths = {}
        for column in sheet3_wb.column_dimensions:
            column_widths[column] = sheet3_wb.column_dimensions[column].width

        # 写入表头
        headers = ['sku编码', 'spu编码', '商品名称', '规格', '第一批时间', '第一批数量', '总交货量'] + sorted_dates
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet3_wb.cell(row=1, column=col_idx)
            cell.value = header

        # 写入数据
        for row_idx, row in enumerate(sheet3_cleaned.values, start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = sheet3_wb.cell(row=row_idx, column=col_idx)
                cell.value = value

        # 复制格式
        # 保存第一行的格式
        first_row_format = []
        for col_idx in range(1, len(headers) + 1):
            cell = sheet3_wb.cell(row=1, column=col_idx)
            first_row_format.append({
                'font': copy(cell.font),
                'border': copy(cell.border),
                'fill': copy(cell.fill),
                'alignment': copy(cell.alignment)
            })

        # 应用保存的格式
        for col_idx in range(1, len(headers) + 1):
            cell = sheet3_wb.cell(row=1, column=col_idx)
            cell.font = first_row_format[col_idx-1]['font']
            cell.border = first_row_format[col_idx-1]['border']
            cell.fill = first_row_format[col_idx-1]['fill']
            cell.alignment = first_row_format[col_idx-1]['alignment']

        # 恢复列宽
        for column, width in column_widths.items():
            sheet3_wb.column_dimensions[column].width = width

        # 在保存文件之前添加格式设置
        from openpyxl.styles import PatternFill

        # 设置颜色
        green_fill = PatternFill(start_color='1F6B3B', end_color='1F6B3B', fill_type='solid')  # 深绿色
        orange_fill = PatternFill(start_color='F4B183', end_color='F4B183', fill_type='solid')  # 浅橙色

        # 获取基础格式单元格(A2)的格式
        base_format_cell = sheet3_wb['A2']
        base_format = {
            'font': copy(base_format_cell.font),
            'border': copy(base_format_cell.border),
            'fill': copy(base_format_cell.fill),
            'alignment': copy(base_format_cell.alignment)
        }

        # 获取H1单元格的格式作为日期列的格式模板
        date_format_cell = sheet3_wb['H1']
        date_format = {
            'font': copy(date_format_cell.font),
            'border': copy(date_format_cell.border),
            'fill': copy(date_format_cell.fill),
            'alignment': copy(date_format_cell.alignment)
        }

        # 应用表头颜色和格式
        fixed_headers = ['sku编码', 'spu编码', '商品名称', '规格', '第一批时间', '第一批数量', '总交货量']
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet3_wb.cell(row=1, column=col_idx)
            if header in fixed_headers:
                cell.fill = green_fill
            else:
                # 对于日期列，使用H1的完整格式
                cell.font = copy(date_format['font'])
                cell.border = copy(date_format['border'])
                cell.fill = copy(date_format['fill'])
                cell.alignment = copy(date_format['alignment'])

        # 设置所有单元格居中对齐
        from openpyxl.styles import Alignment
        center_alignment = Alignment(horizontal='center', vertical='center')

        # 应用A2的格式到所有数据单元格，并设置居中对齐
        max_row = len(sheet3_cleaned) + 1  # +1 因为有表头
        max_col = len(headers)
        for row in range(2, max_row + 1):
            for col in range(1, max_col + 1):
                cell = sheet3_wb.cell(row=row, column=col)
                cell.font = copy(base_format['font'])
                cell.border = copy(base_format['border'])
                cell.fill = copy(base_format['fill'])
                cell.alignment = center_alignment

        # 在处理完数据后添加校验
        def verify_totals(sheet1, sheet2, sheet3_cleaned):
            # 计算常规产品总和
            sheet1_total = 0
            sheet1_sku_totals = {}
            
            # 确保数据类型为数值型
            for idx in range(1, 6):
                qty_col = f'到货数量-{idx}'
                if qty_col in sheet1.columns:
                    # 将数据转换为数值类型
                    sheet1[qty_col] = pd.to_numeric(sheet1[qty_col], errors='coerce')
                    # 使用sum()前先填充NA值为0
                    sheet1_total += sheet1[qty_col].fillna(0).sum()
                    
                    # 计算每个SKU的总量
                    for sku, qty in zip(sheet1['sku编码'], sheet1[qty_col]):
                        if pd.notna(qty):
                            if sku not in sheet1_sku_totals:
                                sheet1_sku_totals[sku] = 0
                            sheet1_sku_totals[sku] += qty
            
            # 计算S级产品总和
            sheet2_total = 0
            sheet2_sku_totals = {}
            date_cols = [col for col in sheet2.columns if pd.notna(pd.to_datetime(col, errors='coerce'))]
            
            for _, row in sheet2.iterrows():
                sku = row['sku编码']
                sku_total = 0
                for col in date_cols:
                    # 确保数据为数值类型
                    qty = pd.to_numeric(row[col], errors='coerce')
                    if pd.notna(qty):
                        sku_total += qty
                        sheet2_total += qty
                if sku_total > 0:
                    sheet2_sku_totals[sku] = sku_total
            
            # 获取汇总表总和
            sheet3_sku_totals = {}
            for _, row in sheet3_cleaned.iterrows():
                sku = row['sku编码']
                total = pd.to_numeric(row['总交货量'], errors='coerce')
                if pd.notna(total) and total > 0:
                    sheet3_sku_totals[sku] = total
            
            sheet3_total = sheet3_cleaned['总交货量'].sum()
            
            # 输出详细信息
            print(f"\n常规产品总和: {sheet1_total:,.0f}")
            print(f"S级产品总和: {sheet2_total:,.0f}")
            print(f"常规+S级总和: {sheet1_total + sheet2_total:,.0f}")
            print(f"汇总表总和: {sheet3_total:,.0f}")
            
            # 检查不一致的SKU
            all_skus = set(sheet1_sku_totals.keys()) | set(sheet2_sku_totals.keys())
            for sku in all_skus:
                original_total = (sheet1_sku_totals.get(sku, 0) + sheet2_sku_totals.get(sku, 0))
                summary_total = sheet3_sku_totals.get(sku, 0)
                if abs(original_total - summary_total) > 0.1:  # 考虑浮点数误差
                    print(f"\nSKU {sku} 数量不一致:")
                    print(f"  常规产品数量: {sheet1_sku_totals.get(sku, 0):,.0f}")
                    print(f"  S级产品数量: {sheet2_sku_totals.get(sku, 0):,.0f}")
                    print(f"  原表总和: {original_total:,.0f}")
                    print(f"  汇总表数量: {summary_total:,.0f}")
            
            return sheet1_total, sheet2_total, sheet3_total

        # 在主处理流程中调用验证
        verify_totals(sheet1, sheet2, sheet3_cleaned)

        # 保存文件
        wb.save(output_path)
        return True, "处理完成"
        
    except Exception as e:
        return False, f"处理失败: {str(e)}"

# 使用示例
input_file = "updated_表优化.xlsx"
success, message = process_excel_file(input_file)
print(message)
