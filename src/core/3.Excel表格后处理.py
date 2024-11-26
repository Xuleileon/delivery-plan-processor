import pandas as pd
from datetime import datetime
import os
import openpyxl
from openpyxl.styles import PatternFill

def format_sku(x):
    """格式化SKU编码，处理科学计数法的问题"""
    if pd.isna(x):
        return ''
    # 如果是数字（包括科学计数法），转换为整数再转字符串
    if isinstance(x, (int, float)):
        return str(int(x))
    return str(x).strip()

def process_excel(input_file):
    try:
        # 生成输出文件路径
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = os.path.splitext(os.path.basename(input_file))[0]
        output_file = f"{filename}_合并_{timestamp}.xlsx"
        output_path = os.path.join(os.path.dirname(input_file), output_file)
        
        # 读取汇总sheet，将所有列都设置为字符串格式
        df = pd.read_excel(input_file, sheet_name='汇总', dtype=str)
        
        # 格式化所有列
        for col in df.columns:
            df[col] = df[col].apply(format_sku)
        
        # 获取日期列
        date_columns = [col for col in df.columns if str(col).startswith('202')]
        
        # 将日期列转换为数值类型进行计算
        for col in date_columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        
        # 显示所有重复的SKU
        duplicates = df[df.duplicated('sku编码', keep=False)].sort_values('sku编码')
        if not duplicates.empty:
            print("\n=== 重复的SKU ===")
            for sku in duplicates['sku编码'].unique():
                print(f"\nSKU: {sku}")
                sku_data = df[df['sku编码'] == sku]
                print(sku_data[['sku编码', '商品名称']].to_string())
                print("-" * 50)
        
        # 合并重复的SKU行
        df_merged = df.groupby('sku编码').agg({
            '商品名称': 'first',
            **{col: 'sum' for col in date_columns}
        }).reset_index()
        
        # 保存结果，只保留汇总sheet
        df_merged.to_excel(output_path, sheet_name='汇总', index=False)
        
        # 使用openpyxl设置单元格格式为文本
        wb = openpyxl.load_workbook(output_path)
        ws = wb['汇总']
        
        # 将所有单元格设置为文本格式
        for row in ws.iter_rows():
            for cell in row:
                cell.number_format = '@'
        
        wb.save(output_path)
        
        print(f"\n处理完成:")
        print(f"原始行数: {len(df)}")
        print(f"处理后行数: {len(df_merged)}")
        print(f"合并的SKU数: {len(duplicates['sku编码'].unique()) if not duplicates.empty else 0}")
        
        return True, f"处理完成,文件保存为: {output_file}"
        
    except Exception as e:
        return False, f"处理失败: {str(e)}"

# 使用示例
if __name__ == "__main__":
    input_file = "updated_表优化_20241121_171723.xlsx"  
    success, message = process_excel(input_file)
    print(message)