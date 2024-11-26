import pytest
import openpyxl
from pathlib import Path
from src.preprocessor.excel_preprocessor import ExcelPreprocessor
from src.utils.excel_utils import setup_logging

class TestExcelPreprocessor:
    @pytest.fixture(scope="class")
    def setup_test_env(self):
        # 创建日志目录
        Path("logs").mkdir(exist_ok=True)
        # 创建测试数据目录
        Path("tests/fixtures").mkdir(parents=True, exist_ok=True)
        Path("output").mkdir(exist_ok=True)
        # 设置日志
        setup_logging()
        
    @pytest.fixture
    def preprocessor(self, setup_test_env):
        return ExcelPreprocessor()
    
    @pytest.fixture
    def sample_excel(self, tmp_path):
        """创建一个测试用的Excel文件"""
        import openpyxl
        wb = openpyxl.Workbook()
        
        # 创建测试工作表
        for sheet_name in ['常规产品', 'S级产品', '汇总']:
            ws = wb.create_sheet(sheet_name)
            ws['A1'] = 'SKU编码'
            ws['B1'] = '数量'
        
        # 删除默认的Sheet
        wb.remove(wb['Sheet'])
        
        # 保存测试文件
        test_file = tmp_path / "sample_excel.xlsx"
        wb.save(test_file)
        return test_file
    
    def test_process_valid_file(self, preprocessor, sample_excel):
        """测试处理有效文件"""
        result = preprocessor.process(sample_excel)
        assert result.exists()
        assert result.suffix == '.xlsx'
    
    def test_process_invalid_file(self, preprocessor):
        """测试处理不存在的文件"""
        with pytest.raises(FileOperationError, match="输入文件不存在"):
            preprocessor.process(Path("not_exists.xlsx"))
    
    @pytest.mark.slow
    def test_large_file_processing(self, preprocessor, tmp_path):
        """测试大文件处理"""
        # TODO: 实现大文件测试
        pass
    
    @pytest.mark.integration
    def test_full_workflow(self, preprocessor, sample_excel):
        """测试完整工作流"""
        # 处理文件
        result = preprocessor.process(sample_excel)
        
        # 验证输出
        assert result.exists()
        
        # 验证输出文件内容
        import openpyxl
        wb = openpyxl.load_workbook(result)
        
        # 验证工作表
        assert set(wb.sheetnames) == {'常规产品', 'S级产品', '汇总'}
        
        # 验证表头
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            assert ws['A1'].value == 'SKU编码'
            assert ws['B1'].value == '数量'
        
        # 打印处理后的文件路径
        print(f"\n处理后的文件保存在: {result.absolute()}")
    
    @pytest.fixture
    def sample_excel_with_data(self, tmp_path):
        """创建一个包含测试数据的Excel文件"""
        import openpyxl
        wb = openpyxl.Workbook()
        
        # 创建测试工作表并添加数据
        data = {
            '常规产品': [
                ['SKU编码', '数量', '日期'],
                ['SKU001', 100, '2024-01-01'],
                ['SKU002', 200, '2024-01-02'],
            ],
            'S级产品': [
                ['SKU编码', '数量', '日期'],
                ['SKU003', 300, '2024-01-01'],
                ['SKU004', 400, '2024-01-02'],
            ],
            '汇总': [
                ['SKU编码', '总数量', '最早日期'],
                ['=常规产品!A2', '=SUM(常规产品!B2:B3)', '=MIN(常规产品!C2:C3)'],
                ['=S级产品!A2', '=SUM(S级产品!B2:B3)', '=MIN(S级产品!C2:C3)'],
            ]
        }
        
        for sheet_name, sheet_data in data.items():
            ws = wb.create_sheet(sheet_name)
            for row in sheet_data:
                ws.append(row)
        
        # 删除默认的Sheet
        wb.remove(wb['Sheet'])
        
        # 保存测试文件
        test_file = tmp_path / "sample_excel_with_data.xlsx"
        wb.save(test_file)
        return test_file
    
    def test_process_with_real_data(self, preprocessor, sample_excel_with_data):
        """测试处理真实数据"""
        result = preprocessor.process(sample_excel_with_data)
        
        # 验证输出
        wb = openpyxl.load_workbook(result)
        
        # 验证汇总表的计算结果
        summary = wb['汇总']
        
        # 打印实际值以便调试
        print(f"\n汇总表数据:")
        for row in summary.iter_rows(min_row=1, max_row=3, values_only=True):
            print(row)
        
        # 验证常规产品总数
        assert summary['B2'].value == 300, f"常规产品总数错误，期望300，实际{summary['B2'].value}"
        
        # 验证S级产品总数
        assert summary['B3'].value == 700, f"S级产品总数错误，期望700，实际{summary['B3'].value}"
        
        # 验证日期
        assert summary['C2'].value == '2024-01-01', f"常规产品最早日期错误，期望2024-01-01，实际{summary['C2'].value}"
        assert summary['C3'].value == '2024-01-01', f"S级产品最早日期错误，期望2024-01-01，实际{summary['C3'].value}"
        
        print(f"\n包含实际数据的处理结果保存在: {result.absolute()}")